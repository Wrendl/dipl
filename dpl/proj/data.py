cars = ['ВАЗ (Lada)', 'ГАЗ', 'ЗАЗ', 'Brilliance', 'BYD', 'Chery', 'Derways', 'FAW', 'Geely', 'Great Wall', 'Hafei', 'Lifan', 'Haima', 'Luxgen', 'Xin Kai', 'Aston Martin', 'Bentley', 'Bugatti', 'Ferrari', 'Jaguar', 'Maserati', 'Porsche', 'Audi', 'BMW', 'Mercedes-Benz', 'Mini', 'Rover', 'Alfa Romeo', 'Citroen', 'Fiat', 'Ford', 'Opel', 'Peugeot', 'Renault', 'Saab', 'SEAT', 'Skoda', 'Volkswagen', 'Volvo', 'Acura', 'Buick', 'Cadillac', 'Chevrolet', 'Chrysler', 'Dodge', 'Hummer', 'Infiniti', 'Jeep', 'Lexus', 'Lincoln', 'Mercury', 'Pontiac', 'Hyundai', 'Kia', 'Ssang Yong', 'Daewoo', 'Daihatsu', 'Honda', 'Isuzu', 'Mazda', 'Mitsubishi', 'Nissan', 'Subaru', 'Suzuki', 'Toyota']
coeff_a = ['0.057', '0.057', '0.057', '0.057', '0.057', '0.057', '0.057', '0.057', '0.057', '0.057', '0.057', '0.057', '0.057', '0.057', '0.057', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.042', '0.045', '0.045', '0.045', '0.045', '0.045', '0.045', '0.045', '0.045', '0.045', '0.045', '0.045', '0.045', '0.045', '0.052', '0.052', '0.052', '0.052', '0.049', '0.049', '0.049', '0.049', '0.049', '0.049', '0.049', '0.049', '0.049']
coeff_b = ['0.003', '0.003', '0.003', '0.0029', '0.0029', '0.0029', '0.0029', '0.0029', '0.0029', '0.0029', '0.0029', '0.0029', '0.0029', '0.0029', '0.0029', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0023', '0.0024', '0.0024', '0.0024', '0.0024', '0.0024', '0.0024', '0.0024', '0.0024', '0.0024', '0.0024', '0.0024', '0.0024', '0.0024', '0.0026', '0.0026', '0.0026', '0.0026', '0.0025', '0.0025', '0.0025', '0.0025', '0.0025', '0.0025', '0.0025', '0.0025', '0.0025']
Lo = ['15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15', '15']
ML = ['0.856', '0.856', '0.856', '0.87', '0.87', '0.87', '0.856', '0.87', '0.87', '0.87', '0.87', '0.87', '0.87', '0.87', '0.87', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.922', '0.895', '0.895', '0.895', '0.895', '0.895', '0.895', '0.922', '0.895', '0.922', '0.895', '0.895', '0.895', '0.87', '0.87', '0.87', '0.87', '0.922', '0.922', '0.922', '0.922', '0.922', '0.922', '0.922', '0.922', '0.922']

import openpyxl


def get_marks_models():
    mark_workbook = openpyxl.load_workbook('../1/list.xlsx', read_only=True)
    mark_sheet = mark_workbook.active
    marks = {}

    for sheet_row in mark_sheet.iter_rows(min_row=2, min_col=1, max_row=mark_sheet.max_row, max_col=9):
        if not sheet_row[0].value in marks:
            marks[sheet_row[0].value] = []

        marks[sheet_row[0].value].append({
                'model_name': sheet_row[1].value,
                'gen_name': sheet_row[2].value,
                'year_begin': sheet_row[3].value,
                'year_end': sheet_row[4].value,
                'coeff_a': sheet_row[5].value,
                'coeff_b': sheet_row[6].value,
                'Lo': sheet_row[7].value,
                'ML': sheet_row[8].value
            }
        )
    # for key, value in marks.items():
    #     print(key, value)
    return marks


def print_models(mark):
    models = get_marks_models()[mark]

    models_list = []
    for model in models:
        str = f'{model["model_name"]}, {model["gen_name"]}, {model["year_begin"]}-{model["year_end"]}'
        models_list.append(str)

    return models_list


def get_models(mark):
    models = get_marks_models()[mark]

    models_list = []
    for model in models:
        models_list.append(model["model_name"])

    return models_list

def get_a_b_L0_ML(mark):
    model = get_marks_models()[mark]

    model_info = {
        'coeff_a': model[0]["coeff_a"],
        'coeff_b': model[0]["coeff_b"],
        'Lo': model[0]["Lo"],
        'ML': model[0]["ML"]
    }

    return model_info


def car_detail():
    workbook = openpyxl.load_workbook('../1/data for calc (1).xlsx', read_only=True)
    sheet = workbook.active
    marks = {}

    for sheet_row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=3):
        if not sheet_row[0].value in marks:
            marks[sheet_row[0].value] = []

        marks[sheet_row[0].value].append({
                'model_name': sheet_row[1].value,
                'gen_name': sheet_row[2].value,
                'year_begin': sheet_row[3].value,
                'year_end': sheet_row[4].value,
                'coeff_a': sheet_row[5].value,
                'coeff_b': sheet_row[6].value,
                'Lo': sheet_row[7].value,
                'ML': sheet_row[8].value
            }
        )
    # for key, value in marks.items():
    #     print(key, value)
    return marks
