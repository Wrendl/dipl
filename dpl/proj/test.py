import requests
from bs4 import BeautifulSoup
import re


def test1():
    import openpyxl

    mark_workbook = openpyxl.load_workbook('../1/list.xlsx', read_only=True)

    mark_sheet = mark_workbook.active

    marks = {}

    for sheet_row in mark_sheet.iter_rows(min_row=2, min_col=1, max_row=mark_sheet.max_row, max_col=5):
        if not sheet_row[0].value in marks:
            marks[sheet_row[0].value] = []

        marks[sheet_row[0].value].append({
                'model_name': sheet_row[1].value,
                'gen_name': sheet_row[2].value,
                'year_begin': sheet_row[3].value,
                'year_end': sheet_row[4].value
            }
        )
    for key, value in marks.items():
        print(key, value)


def aster(miles, prices):
    volume = 1.6
    year = 2022
    brand = 'Kia'
    model = 'Ceed'
    fuel_type = 'Бензин'
    kpp_type = 'Автомат'
    sweel_type = 'Слева'
    privod_type = 'Передний'

    car_transm_st = {'механика': 'MT', 'автомат': 'AVTOMAT',
                     'вариатор': 'VARIATOR', 'робот': 'ROBOT'}
    sweel_st = {'слева': 1, 'справа': 2}
    dwheel_st = {'передний привод': 'front', 'передний': 'front',
                 'полный привод': 'full', 'полный': 'full',
                 'задний привод': 'back', 'задний': 'back'}

    host = f'https://aster.kz/cars/{brand.lower()}/{model.lower()}?yearFrom={year}&yearTo={year}&transmission={car_transm_st[kpp_type.lower()]}&transmissionDriveType={dwheel_st[privod_type.lower()]}&mileageTo={miles}&volumeFrom={volume}&volumeTo={volume}&steering={sweel_st[sweel_type.lower()]}'

    request = requests.get(host)
    soup = BeautifulSoup(request.text, 'html.parser')

    items_soup = soup.find_all(class_='car fadeIn')
    for item in items_soup:
        # print(item.find('a', class_='price fw-700 f-18 car-link').text)
        try:
            item_price = re.sub("[^0-9]", "", item.find('a', class_='px-3 pt-1 price fw-700 f-18 car-link').text)
            prices.append(int(item_price))
        except:
            continue

    return prices


if __name__ == '__main__':
    # test1()
    prices = []
    print(aster(10000, prices))