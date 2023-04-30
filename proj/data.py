import os
import openpyxl
from datetime import datetime


def get_marks_models():
    mark_workbook = openpyxl.load_workbook('../1/list.xlsx', read_only=True)
    mark_sheet = mark_workbook["Марка"]
    marks = {}

    for sheet_row in mark_sheet.iter_rows(min_row=2, min_col=1, max_row=mark_sheet.max_row, max_col=9):
        if not sheet_row[0].value in marks:
            marks[sheet_row[0].value] = []

        marks[sheet_row[0].value].append({
            'model_name': sheet_row[1].value,
            'gen_name': sheet_row[2].value,
            'year_begin': sheet_row[3].value,
            'year_end': sheet_row[4].value,
            'coeff_a': sheet_row[5].value,
            'coeff_b': sheet_row[6].value,
            'Lo': sheet_row[7].value,
            'ML': sheet_row[8].value
        }
        )
    # for key, value in marks.items():
    #     print(key, value)
    return marks


def get_class_auto():
    class_workbook = openpyxl.load_workbook('../1/list.xlsx', read_only=True)
    class_sheet = class_workbook["Класс авто"]
    class_auto = {}

    for sheet_row in class_sheet.iter_rows(min_row=3, min_col=1, max_row=class_sheet.max_row, max_col=3):
        if not sheet_row[0].value in class_auto:
            class_auto[sheet_row[0].value] = []

        class_auto[sheet_row[0].value].append(sheet_row[1].value)
        class_auto[sheet_row[0].value].append(sheet_row[2].value)
    return class_auto


def print_models(mark):
    models = get_marks_models()[mark]

    models_list = []
    for model in models:
        str = f'{model["model_name"]}, {model["gen_name"]}, {model["year_begin"]}-{model["year_end"]}'
        models_list.append(str)

    return models_list


def get_a_b_L0_ML(mark):
    model = get_marks_models()[mark]

    model_info = {
        'coeff_a': model[0]["coeff_a"],
        'coeff_b': model[0]["coeff_b"],
        'Lo': model[0]["Lo"],
        'ML': model[0]["ML"]
    }

    return model_info


def get_car_details():
    detail_workbook = openpyxl.load_workbook('../1/list.xlsx', read_only=True)
    detail_sheet = detail_workbook["Детали"]
    detail_auto = {}

    for sheet_row in detail_sheet.iter_rows(min_row=2, min_col=1, max_row=detail_sheet.max_row, max_col=2):
        if not sheet_row[0].value in detail_auto:
            detail_auto[sheet_row[0].value] = []

        detail_auto[sheet_row[0].value].append(sheet_row[1].value)
    return detail_auto


def get_material_price():
    price_workbook = openpyxl.load_workbook('../1/list.xlsx', read_only=True)
    price_sheet = price_workbook["Цена материалов"]

    total_price = 0
    for sheet_row in price_sheet.iter_rows(min_row=1, min_col=2, max_row=price_sheet.max_row, max_col=3):
        total_price += sheet_row[0].value * sheet_row[1].value
    return total_price


def get_norm_chas(class_auto, year):
    norm_chas_workbook = openpyxl.load_workbook('../1/list.xlsx', read_only=True)
    norm_chas_sheet = norm_chas_workbook["Норм часы по классам"]
    mrp = norm_chas_sheet["D2"].value

    for sheet_row in norm_chas_sheet.iter_rows(min_row=3, min_col=1, max_row=norm_chas_sheet.max_row, max_col=3):
        if sheet_row[0].value == class_auto:
            if (year - datetime.now().year) < 5:
                return sheet_row[1].value * mrp
            return sheet_row[2].value * mrp


def data_to_excel(arr, wb, sheet_type, header):
    if sheet_type not in wb.sheetnames:
        ws = wb.create_sheet(sheet_type)
        ws.append(header)

    ws = wb[sheet_type]
    ws.append(arr)


def close_excel(wb):
    wb.remove(wb['Sheet'])
    file_name = fr"output\{datetime.now().strftime('%d.%m.%Y_%H.%M.%S')}.xlsx"
    dir = os.path.abspath('..')
    file_path = os.path.join(dir, file_name)
    wb.save(file_path)
    return file_path
